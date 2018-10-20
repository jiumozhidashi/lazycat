# from django.http import HttpResponse
from django.shortcuts import render
from django.http import FileResponse
from django.utils.encoding import escape_uri_path
from django.http import HttpResponse
import pandas as pd
# import numpy as np
# import openpyxl
# import difflib
from openpyxl import load_workbook
# from itertools import islice


def hello(request):
    context = {}
    context['hello'] = 'Hello World!'
    return render(request, 'hello.html', context)


def formatysb(request):
    # context = {}
    # context['ysb'] = 'Hello World!'
    return render(request, 'formatysb.html')


def downloadysbmodel():
    file = open('static/downloads/演算表模板.xlsx', 'rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path('演算表模板.xlsx'))
    return response


def uploadfile(request):
    global obj1, os, s
    if request.method == 'POST':  # 获取对象
        obj1 = request.FILES.get('calc_wb')
        obj2 = request.FILES.get('task_wb')
        obj3 = request.FILES.get('mis_wb')
        import os  # 上传文件的文件名
    print(obj1.name)
    f = open(os.path.join('static', 'downloads', obj1.name), 'wb')
    for chunk in obj1.chunks():
        f.write(chunk)
    f.close()
    wb = load_workbook('static/downloads/%s' % obj1.name)
    oldws = wb.active
    station_row = 4
    station_start_col = 7
    material_start_row = 7
    materialinfo_start_col = 2
    materialinfo_end_col = 5
    # materialqty_start_col = 7
    for cell in oldws[station_row]:
        if cell.value == "单价(元)":
            break
        elif cell.value is not None:
            s = cell.value
        else:
            cell.value = s
    newws = wb.create_sheet(title='formatsheet')
    header = ('物料名称', '规格', '单位', '数量', '工程名称')
    c = 1
    for head in header:
        newws.cell(row=1,column=c,value=head)
        c += 1
    write_row = 2
    read_max_row = 143
    read_col = station_start_col
    for st_col in oldws.iter_cols(min_row=station_row, max_row=station_row, min_col=station_start_col):
        for station_cell in st_col:
            if station_cell.value == "单价(元)" or station_cell.value is None:
                break
            else:
                read_row = material_start_row
                for ma_row in oldws.iter_rows(min_row=material_start_row, max_row=read_max_row,
                                              min_col=materialinfo_start_col,
                                              max_col=materialinfo_end_col):
                    write_col = 1
                    material_qty=oldws.cell(row=read_row,column=read_col).value
                    if (ma_row[3].value == '甲供')and (material_qty != 0) and (material_qty is not None):
                        for r in ma_row[0:3]:
                            newws.cell(row=write_row, column=write_col, value=r.value)
                            write_col += 1
                        newws.cell(row=write_row, column=write_col, value=material_qty)
                        newws.cell(row=write_row, column=write_col + 1, value=station_cell.value)
                        write_row += 1
                    read_row += 1
                read_col += 1
    wb.save('static/downloads/%s' % obj1.name)
    df_ysb=pd.read_excel('static/downloads/%s' % obj1.name,sheet_name='formatsheet')
    f2 = open(os.path.join('static', 'downloads', obj2.name), 'wb')
    for chunk in obj2.chunks():
        f2.write(chunk)
    f2.close()
    df_task_list=pd.read_excel('static/downloads/%s' % obj2.name)
    df_material_use=pd.merge(df_ysb,df_task_list[['工程名称','MIS任务号']],on='工程名称',how='outer')
    df_material_use=df_material_use.round({'数量':3})
    df_material_name=df_material_use['物料名称'].str.cat(df_material_use['规格'], na_rep='').str.strip()
    df_material_use=pd.concat([df_material_name,df_material_use[['单位','数量','工程名称','MIS任务号']]],axis=1)
    print(df_material_use)
    df_material_name=df_material_name.drop_duplicates(keep='first', inplace=False)
    df_material_name=df_material_name.reset_index()['物料名称']
    print(df_material_name)
    f3 = open(os.path.join('static', 'downloads', obj3.name), 'wb')
    for chunk in obj3.chunks():
        f3.write(chunk)
    f3.close()
    df_mis = pd.read_excel('static/downloads/%s' % obj3.name)
    df_material_mis_name=df_mis['备注']
    df_material_mis_name=df_material_mis_name.drop_duplicates(keep='first', inplace=False)
    df_material_mis_name=df_material_mis_name.reset_index()['备注']
    df_rules = pd.read_excel('static/rules.xlsx')
    df_mis=pd.merge(df_mis,df_rules,left_on='备注',right_on='MIS物料名称',how='left')
    df_mis_grouped=df_mis.groupby(['演算表物料名称','任务'],as_index=False).sum()
    df_material_use_grouped=df_material_use.groupby(['物料名称','MIS任务号'],as_index=False).sum()
    df_material_use_grouped['数量']=-df_material_use_grouped['数量']
    print(df_mis_grouped)
    print(df_material_use_grouped)
    #df_merge = pd.merge(df_mis, df_material_use, left_on=['演算表物料名称', '任务'], right_on=['物料名称', 'MIS任务号'],
      #                  how='outer')
    df_mis_grouped.to_excel("static/df_mis_grouped.xlsx", sheet_name="01", index=False, header=True)
    return HttpResponse('OK')
    # return render(request, 'formatysb.html')
